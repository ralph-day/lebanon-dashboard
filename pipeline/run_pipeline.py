#!/usr/bin/env python3
"""Lebanon 2026 survey pipeline — replaces the Excel/VBS refresh loop.

Every run:
  1. pulls all submissions for the form from the SurveyCTO REST API
  2. reads the GTS Google Sheet (manual updates by Moe) via the Drive API
  3. rebuilds all 9 query outputs in pandas (transforms.py)
  4. rewrites the output workbook's query sheets + Dashboard values,
     leaving Target / transferLoc / Rules untouched
  5. uploads the workbook back to the shared Drive folder (same file id,
     so the dashboard's newest-file lookup keeps working)

Designed to finish in seconds and to run every 5 minutes from a cloud
scheduler (e.g. Railway cron) — no Excel, no VBS, no laptop required.

Environment (or .env next to this script):
  SURVEYCTO_SERVER             SurveyCTO server subdomain, e.g. 'gts'
  SURVEYCTO_FORM_ID            form id, e.g. 'lebanon_2026'
  SURVEYCTO_USER               SurveyCTO account email (API access enabled)
  SURVEYCTO_PASSWORD           SurveyCTO account password
  GOOGLE_SERVICE_ACCOUNT_JSON  path to the service-account key file
  DRIVE_FOLDER_ID              shared folder id (same one the dashboard reads)
  OUTPUT_FILE_NAME             default: Lebanon 2026 - Analysis.xlsx
  GTS_SHEET_NAME                default: GTS Master sheet

Offline testing:
  python run_pipeline.py --offline WIDE.xlsx GTS.xlsx WORKBOOK.xlsx OUT.xlsx
"""
from __future__ import annotations

import io
import os
import sys
import time
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

import transforms

OUTPUT_FILE_NAME = os.environ.get("OUTPUT_FILE_NAME", "Lebanon 2026 - Analysis.xlsx")
GTS_SHEET_NAME = os.environ.get("GTS_SHEET_NAME", "GTS Master sheet")
QUERY_SHEETS = ["data", "GTS DATA", "QA_TimingSections", "QA_Dashboard", "QA_ByGroupSection",
                "Query_All_Rules", "Enumerator_Summary", "Target_Tracker", "Survey Comparison"]


# ──────────────────────────────────────────────────────────────────
# Google Drive I/O (service account — same key the dashboard uses)
# ──────────────────────────────────────────────────────────────────

def fetch_wide_from_surveycto() -> pd.DataFrame:
    """Pull all submissions for the form via the SurveyCTO REST API —
    replaces the local `Lebanon 2026_WIDE.xlsx` export so the pipeline
    no longer depends on a laptop running SurveyCTO Desktop sync.

    Environment:
      SURVEYCTO_SERVER     server subdomain, e.g. 'gts'
      SURVEYCTO_FORM_ID    form id, e.g. 'lebanon_2026'
      SURVEYCTO_USER       account email (must have API data access)
      SURVEYCTO_PASSWORD   account password
    """
    import requests

    server = os.environ["SURVEYCTO_SERVER"]
    form_id = os.environ["SURVEYCTO_FORM_ID"]
    user = os.environ["SURVEYCTO_USER"]
    password = os.environ["SURVEYCTO_PASSWORD"]

    url = f"https://{server}.surveycto.com/api/v2/forms/data/wide/json/{form_id}"
    resp = requests.get(url, params={"date": "0"}, auth=(user, password), timeout=120)
    resp.raise_for_status()
    rows = resp.json()
    return pd.DataFrame(rows)


def drive_client():
    from google.oauth2 import service_account
    from googleapiclient.discovery import build

    key_path = os.environ["GOOGLE_SERVICE_ACCOUNT_JSON"]
    creds = service_account.Credentials.from_service_account_file(
        key_path, scopes=["https://www.googleapis.com/auth/drive"])
    return build("drive", "v3", credentials=creds, cache_discovery=False)


def find_file(drive, folder_id: str, name: str, mime_prefix: str | None = None) -> dict:
    q = f"'{folder_id}' in parents and name = '{name}' and trashed = false"
    files = drive.files().list(q=q, fields="files(id, name, mimeType, modifiedTime)").execute()["files"]
    if mime_prefix:
        files = [f for f in files if f["mimeType"].startswith(mime_prefix)]
    if not files:
        raise FileNotFoundError(f"'{name}' not found in Drive folder {folder_id}")
    return files[0]


def download_xlsx(drive, file: dict) -> bytes:
    from googleapiclient.http import MediaIoBaseDownload

    if file["mimeType"] == "application/vnd.google-apps.spreadsheet":
        # Native Google Sheet → export as xlsx
        request = drive.files().export_media(
            fileId=file["id"],
            mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    else:
        request = drive.files().get_media(fileId=file["id"])
    buf = io.BytesIO()
    downloader = MediaIoBaseDownload(buf, request)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    return buf.getvalue()


def upload_xlsx(drive, file_id: str, payload: bytes):
    from googleapiclient.http import MediaIoBaseUpload

    media = MediaIoBaseUpload(
        io.BytesIO(payload),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        resumable=True)
    drive.files().update(fileId=file_id, media_body=media).execute()


# ──────────────────────────────────────────────────────────────────
# Workbook writer
# ──────────────────────────────────────────────────────────────────

def _write_frame(ws, df: pd.DataFrame):
    ws.delete_rows(1, ws.max_row)
    ws.append(list(df.columns))
    for row in df.itertuples(index=False, name=None):
        ws.append([None if (isinstance(v, float) and pd.isna(v)) or v is pd.NaT or v is None
                   else (v.to_pydatetime() if isinstance(v, pd.Timestamp) else v)
                   for v in row])


def _update_dashboard(ws, frames: dict[str, pd.DataFrame]):
    """Replace the Dashboard sheet's formula cells with computed values —
    the dashboard server reads cached values, which would otherwise go
    stale once Excel stops recalculating the file."""
    tracker = frames["Target_Tracker"]
    qa = frames["QA_Dashboard"]
    today = pd.Timestamp.now().normalize()
    submissions = pd.to_datetime(qa["SubmissionDate"], errors="coerce")

    total_target = int(tracker["target"].sum())
    accepted = int(tracker["Accepted"].sum())

    values = {
        "B6": int(tracker["location"].notna().sum()),                       # Total areas
        "E6": total_target,                                                  # Total target
        "H6": total_target - accepted,                                       # Remaining
        "K6": int(((submissions >= today) & (submissions < today + pd.Timedelta(days=1))).sum()),
        "E9": accepted,                                                      # Completed (accepted)
        "H9": (accepted / total_target) if total_target else 0,              # % Complete
        "E11": int(tracker["Completed"].sum()),                              # total completed
        "B14": int(tracker["Palestinian"].sum()),
        "E14": int(tracker["Syrian"].sum()),
        "H14": int(tracker["Lebanese"].sum()),
        "B22": int(tracker["woman"].sum()),
        "E22": int(tracker["man"].sum()),
    }
    values["B17"] = values["B14"] / 200
    values["E17"] = values["E14"] / 200
    values["H17"] = values["H14"] / 800
    for ref, value in values.items():
        ws[ref] = value


def rebuild_workbook(workbook_bytes: bytes, frames: dict[str, pd.DataFrame]) -> bytes:
    wb = load_workbook(io.BytesIO(workbook_bytes))
    for sheet_name in QUERY_SHEETS:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"output workbook is missing sheet '{sheet_name}'")
        _write_frame(wb[sheet_name], frames[sheet_name])
    if "Dashboard" in wb.sheetnames:
        _update_dashboard(wb["Dashboard"], frames)
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ──────────────────────────────────────────────────────────────────
# Main
# ──────────────────────────────────────────────────────────────────

def run_offline(wide_path, gts_path, workbook_path, out_path):
    wide = pd.read_excel(wide_path, sheet_name="data")
    gts = pd.read_excel(gts_path, sheet_name="Query1")
    workbook_bytes = Path(workbook_path).read_bytes()
    target = pd.read_excel(io.BytesIO(workbook_bytes), sheet_name="Target")

    frames = transforms.run_all(wide, gts, target)
    Path(out_path).write_bytes(rebuild_workbook(workbook_bytes, frames))
    print(f"offline run OK → {out_path}")
    for name, df in frames.items():
        print(f"  {name}: {len(df)} rows x {len(df.columns)} cols")


def run_live():
    started = time.time()
    folder_id = os.environ["DRIVE_FOLDER_ID"]

    wide = fetch_wide_from_surveycto()
    print(f"[{datetime.now():%H:%M:%S}] WIDE: {len(wide)} rows from SurveyCTO API")

    drive = drive_client()
    gts_file = find_file(drive, folder_id, GTS_SHEET_NAME)
    gts = pd.read_excel(io.BytesIO(download_xlsx(drive, gts_file)), sheet_name="Query1")
    print(f"[{datetime.now():%H:%M:%S}] GTS: {len(gts)} rows ({gts_file['mimeType']})")

    out_file = find_file(drive, folder_id, OUTPUT_FILE_NAME, mime_prefix="application/vnd.openxmlformats")
    workbook_bytes = download_xlsx(drive, out_file)
    target = pd.read_excel(io.BytesIO(workbook_bytes), sheet_name="Target")

    frames = transforms.run_all(wide, gts, target)
    payload = rebuild_workbook(workbook_bytes, frames)
    upload_xlsx(drive, out_file["id"], payload)
    print(f"[{datetime.now():%H:%M:%S}] uploaded {len(payload) // 1024} KB "
          f"({frames['data'].shape[0]} surveys) in {time.time() - started:.1f}s")


if __name__ == "__main__":
    # Optional .env loader (no dependency if python-dotenv is absent)
    env_file = Path(__file__).parent / ".env"
    if env_file.exists():
        for line in env_file.read_text().splitlines():
            if line.strip() and not line.startswith("#") and "=" in line:
                k, v = line.split("=", 1)
                os.environ.setdefault(k.strip(), v.strip())

    if len(sys.argv) > 1 and sys.argv[1] == "--offline":
        run_offline(*sys.argv[2:6])
    else:
        run_live()
